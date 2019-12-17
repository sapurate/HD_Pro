
#查询 5分钟内正在执行任务的话加团队外呼效能
import os
import pymysql
import time
import prettytable as pt
import pandas as pd
import copy
import datetime
import openpyxl
#此外关于中文字串前加“u”的问题，是 python 在 2.x 版本的一个历史性缺陷，只要加上

# 定义全局变量
db_connection = pymysql.connect(
    host='172.42.28.19',
    port=8066,
    user='hdzf',
    password='#123zhangf',
    database='huaplus',
    charset='utf8'
    )

h_sql_1 = '''SELECT
 project_record.agent_id,
 agent.company_name,
 project_record.batch_id,
 batch.name,
 COUNT(project_record.called_num),
 COUNT(IF(project_record.call_duration > 0,true,null)) ,
 COUNT(DISTINCT IF(project_record.end_result=4,project_record.called_num,null)),
 (COUNT(DISTINCT IF(project_record.end_result=4,project_record.called_num,null)))/(COUNT(IF(project_record.call_duration > 0,true,null))),
 (COUNT(IF(project_record.call_duration > 0,true,null)))/(COUNT(project_record.called_num)),
 COUNT(DISTINCT project_record.uid)
FROM
 project_record,batch,agent
WHERE
project_record.agent_id = agent.id
AND
project_record.batch_id = batch.id
AND
project_record.agent_id IN (
'1238','1239','1461815','1461770','1461616'
)
AND project_record.ctime BETWEEN
'''

h_sql_5min_2 = '''
  GROUP BY project_record.agent_id,project_record.batch_id
'''
def ConnectToData(db_connection):
    try:
        cur = db_connection.cursor()
        return cur
    except Exception as e:
        print(e)

def GetData(sql,cursors):
    try:
        cursors.execute(sql)
        results = cursors.fetchall()
        cursors.close()
        return results
    except Exception as e:
        print(e)
def CreateWb(filepath,wbname,sheetname):
    #表名存在就创建，存在就跳过
    path = os.path.join(filepath + "\\" +wbname + ".xls")
    p1 = os.path.exists(path)
    if p1:
        pass # 如果存在就不管
    else:    # 如果不存在就创建
        wb=openpyxl.Workbook()
        wb.save(filename=path.replace('\\','/'))
    #获得已存在表的对象wb
    wb=openpyxl.load_workbook(filename=path.replace('\\','/'))
    #根据表名sheetname，删除旧表，创建新表
    if sheetname in wb.sheetnames:
        del wb[sheetname] #删除旧表,目前只能用del,remove还不行
        ws = wb.create_sheet(sheetname) #新建空表，得到表对象ws
    else:
        ws = wb.create_sheet(sheetname) #新建空表，得到表对象ws
    #准备表头
    # 数据写入是按照1,1的方式 range 计数是0、1、2....6,并初始化为0
    wb.save(filename=path.replace('\\','/'))
    return wb,ws #返回可操作的对象: 整个工作簿wb,相应工作表ws
#if __name__ == '__main__':
while(1):
#    bef_time = time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(time.time()-900))
    now_time = time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(time.time()))
    # 今天最小时间（今天0点）
    today_minTime01 = datetime.datetime.combine(datetime.datetime.now(),datetime.time.min).strftime('%Y-%m-%d %H:%M:%S')
    # 今天最大时间（24点）
    #today_maxTime02 = datetime.datetime.combine(datetime.datetime.now(),datetime.time.max).strftime('%Y-%m-%d %H:%M:%S')
#     start_time="2019-10-12 00:00:00"
#     end_time  ="2019-10-12 09:10:14"

    start_time = today_minTime01
    end_time = now_time
    #h_sql_1=h_sql_5min_1%("'1238','1239','1461815','1461770','1461616'")
    #print(h_sql_1)
    QLSQL = h_sql_1+'''"'''+start_time+ '''"'''+" AND " +'''"'''+ end_time +'''"'''+ h_sql_5min_2
    #print(QLSQL)
    #tb = pt.PrettyTable()
    #tb.field_names= (["渠道id", "渠道名","批次id", "批次名", "外呼量","接通量","成功量","成功率","接通率","上线人数"])

    QLResult = GetData(QLSQL, ConnectToData(db_connection))
    #for i in range(0,len(QLResult)):
    #    tb.add_row(QLResult[i])
    #print(tb)
    #db_connection.close() #查完即关闭

    #定义临时变量
    a = []
    b = []
    for i in range(0,len(QLResult)):
        for j in range(0,len(QLResult[i])):
            a.append(QLResult[i][j])
        # 使用深拷贝不影响a改变 同时改变b的尴尬
        b.append(copy.deepcopy(a))
#         print(b)
        a.clear()
#     print(b)
    df = pd.DataFrame(b,columns=["渠道id", "渠道名","批次id", "%s    批次名"%end_time, "外呼量","接通量","成功量","成功率","接通率","上线人数"])
    df.to_html(r"\\Hll\湖北话加\min_data.html")
    os.system(""">> //Hll/湖北话加/min_data.html echo ^<meta http-equiv='refresh' content='10'^>""")
    print("%s刷新完毕"%end_time)
#     read_html = pd.read_html("min_data.html")
#     start_time1='2019-10-12 00:00:00'
#     CreateWb(r"D:\Users\Administrator\workspace",'excel_output.xls',end_time)
#    df.to_excel('excel_output.xls',sheet_name='Sheet1')
    #os.startfile("min_data.html")
    time.sleep(60)